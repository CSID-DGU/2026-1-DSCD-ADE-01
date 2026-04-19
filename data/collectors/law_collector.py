import re
import requests
import xml.etree.ElementTree as ET
from pathlib import Path
from openpyxl import Workbook
import os
from dotenv import load_dotenv

load_dotenv(Path(__file__).parent.parent.parent / ".env")
OC = os.getenv("LAW_API_KEY")
BASE_URL = "https://www.law.go.kr/DRF"

# 항내용 앞 원기호(①②③...) 제거 패턴
CIRCLE_RE = re.compile(r"^[①-⑳㉑-㉟]+\s*")
# 호/목 내용 앞 중복 번호("1.  ", "가.  " 등) 제거 패턴
HO_NUM_RE = re.compile(r"^[\d가-힣]+\.\s*")

# ── 수집 대상 ─────────────────────────────────────────────────────────

LAW_NAMES = [
    # 핵심 11개 법령
    "주택임대차보호법",
    "민법",
    "부동산등기법",
    "공인중개사법",
    "화물자동차 운수사업법",
    "민사소송법",
    "민사조정법",
    "민사집행법",
    "소액사건심판법",
    "공동주택관리법",
    "주민등록법",
    # 판례 참조 법령 추가
    "채무자 회생 및 파산에 관한 법률",
    "국토의 계획 및 이용에 관한 법률",
    "건축법",
    "건설기술 진흥법",
    "형사소송법",
    "지방세기본법",
    "국가를 당사자로 하는 계약에 관한 법률",
    "약관의 규제에 관한 법률",
    "농지법",
    "상속세 및 증여세법",
    "도시 및 주거환경정비법",
    "건축물의 분양에 관한 법률",
    "상고심절차에 관한 특례법",
    "형법",
    "지방세법",
    "신탁법",
    "국유재산법",
    "소송촉진 등에 관한 특례법",
    "상법",
    "공공기관의 운영에 관한 법률",
    "하도급거래 공정화에 관한 법률",
]

ENFORCEMENT_DECREE_NAMES = [
    # 기존 법률 시행령
    "주택임대차보호법 시행령",
    "공인중개사법 시행령",
    "화물자동차 운수사업법 시행령",
    "민사집행법 시행령",
    "공동주택관리법 시행령",
    "주민등록법 시행령",
    # 판례 참조 시행령
    "농지법 시행령",
    "국가를 당사자로 하는 계약에 관한 법률 시행령",
    "건축법 시행령",
    "지방세기본법 시행령",
    "국토의 계획 및 이용에 관한 법률 시행령",
]

ENFORCEMENT_RULE_NAMES = [
    "농지법 시행규칙",
    "국가를 당사자로 하는 계약에 관한 법률 시행규칙",
]


# ── 검색 함수 ─────────────────────────────────────────────────────────

def search_law(query: str, law_type_filter: str | None = "법률") -> list[dict]:
    """국가법령정보 API에서 법령명으로 검색한다."""
    url = f"{BASE_URL}/lawSearch.do"
    params = {
        "OC": OC,
        "target": "law",
        "type": "XML",
        "query": query,
        "display": 100,
        "page": 1,
    }
    resp = requests.get(url, params=params)
    resp.raise_for_status()
    resp.encoding = "utf-8"

    root = ET.fromstring(resp.text)
    query_normalized = query.replace(" ", "")

    # 1차: 법령명 완전일치 + 법종구분 일치
    for law in root.findall("law"):
        law_name = law.findtext("법령명한글", "").strip()
        법종 = law.findtext("법령구분명", "").strip()
        if law_name.replace(" ", "") == query_normalized:
            if law_type_filter is None or 법종 == law_type_filter:
                return [{
                    "법령명": law_name,
                    "MST": law.findtext("법령일련번호", ""),
                    "시행일": law.findtext("시행일자", ""),
                    "법령구분": 법종,
                }]

    # 2차: 법종구분 무시하고 이름만 일치
    for law in root.findall("law"):
        law_name = law.findtext("법령명한글", "").strip()
        if law_name.replace(" ", "") == query_normalized:
            법종 = law.findtext("법령구분명", "").strip()
            return [{
                "법령명": law_name,
                "MST": law.findtext("법령일련번호", ""),
                "시행일": law.findtext("시행일자", ""),
                "법령구분": 법종,
            }]

    return []


def get_law_text(mst: str) -> str:
    """법령 일련번호(MST)로 법령 원문 XML을 가져온다."""
    url = f"{BASE_URL}/lawService.do"
    params = {
        "OC": OC,
        "target": "law",
        "type": "XML",
        "MST": mst,
    }
    resp = requests.get(url, params=params)
    resp.raise_for_status()
    resp.encoding = "utf-8"
    return resp.text


# ── 파싱 함수 ─────────────────────────────────────────────────────────

def parse_law(xml_text: str):
    """법령 XML을 파싱해서 기본정보, 조문, 항호목, 부칙, 개정이유를 반환한다."""
    root = ET.fromstring(xml_text)

    # 기본정보
    basic = {
        "법령명_한글": root.findtext(".//법령명_한글", "").strip(),
        "법령명_한자": root.findtext(".//법령명_한자", "").strip(),
        "법령명_약칭": root.findtext(".//법령명약칭", "").strip(),
        "법령ID": root.findtext(".//법령ID", "").strip(),
        "법종구분": root.findtext(".//법종구분", "").strip(),
        "소관부처": root.findtext(".//소관부처", "").strip(),
        "소관부처코드": "",
        "공포일자": root.findtext(".//공포일자", "").strip(),
        "공포번호": root.findtext(".//공포번호", "").strip(),
        "시행일자": root.findtext(".//시행일자", "").strip(),
        "제개정구분": root.findtext(".//제개정구분", "").strip(),
        "한글법령여부": root.findtext(".//한글법령여부", "").strip(),
        "별표편집여부": root.findtext(".//별표편집여부", "").strip(),
        "공포법령여부": root.findtext(".//공포법령여부", "").strip(),
        "편장절관": root.findtext(".//편장절관", "").strip(),
        "전화번호": root.findtext(".//부서연락처", "").strip(),
    }
    부처el = root.find(".//소관부처")
    if 부처el is not None:
        basic["소관부처코드"] = 부처el.get("소관부처코드", "")

    법령명 = basic["법령명_한글"]

    # 조문
    articles = []
    for 단위 in root.iter("조문단위"):
        jo_num = 단위.findtext("조문번호", "").strip()
        jo_sub = 단위.findtext("조문가지번호", "").strip()
        display = f"{jo_num}의{jo_sub}" if jo_sub else jo_num
        articles.append({
            "법령명": 법령명,
            "조문번호": display,
            "조문구분": 단위.findtext("조문여부", "").strip(),
            "조문제목": 단위.findtext("조문제목", "").strip(),
            "조문내용": 단위.findtext("조문내용", "").strip(),
            "조문시행일자": 단위.findtext("조문시행일자", "").strip(),
            "조문변경여부": 단위.findtext("조문변경여부", "").strip(),
            "조문참고자료": 단위.findtext("조문참고자료", "").strip(),
        })

    # 항호목
    hang_rows = []
    for 단위 in root.iter("조문단위"):
        jo_num = 단위.findtext("조문번호", "").strip()
        jo_sub = 단위.findtext("조문가지번호", "").strip()
        display = f"{jo_num}의{jo_sub}" if jo_sub else jo_num

        항_list = 단위.findall("항")
        if 항_list:
            for 항 in 항_list:
                항번호 = 항.findtext("항번호", "").strip()
                항내용 = CIRCLE_RE.sub("", 항.findtext("항내용", "").strip())
                항개정 = 항.findtext("항제개정유형", "").strip()
                항날짜 = 항.findtext("항제개정일자", "").strip()

                호목_lines = []
                for 호 in 항.findall("호"):
                    호번호 = 호.findtext("호번호", "").strip().rstrip(".")
                    호내용 = HO_NUM_RE.sub("", 호.findtext("호내용", "").strip())  # 중복 번호 제거
                    호목_lines.append(f"{호번호}. {호내용}")
                    for 목 in 호.findall("목"):
                        목번호 = 목.findtext("목번호", "").strip().rstrip(".")
                        목내용 = HO_NUM_RE.sub("", 목.findtext("목내용", "").strip())  # 중복 번호 제거
                        호목_lines.append(f"  {목번호}. {목내용}")

                hang_rows.append({
                    "법령명": 법령명,
                    "조문번호": display,
                    "항번호": 항번호,
                    "항내용": 항내용,
                    "항제개정유형": 항개정,
                    "항제개정일자": 항날짜,
                    "호_목_목록": "\n".join(호목_lines) if 호목_lines else None,
                })
        else:
            호목_lines = []
            for 호 in 단위.findall("호"):
                호번호 = 호.findtext("호번호", "").strip().rstrip(".")
                호내용 = HO_NUM_RE.sub("", 호.findtext("호내용", "").strip())  # 중복 번호 제거
                호목_lines.append(f"{호번호}. {호내용}")
                for 목 in 호.findall("목"):
                    목번호 = 목.findtext("목번호", "").strip().rstrip(".")
                    목내용 = HO_NUM_RE.sub("", 목.findtext("목내용", "").strip())  # 중복 번호 제거
                    호목_lines.append(f"  {목번호}. {목내용}")
            if 호목_lines:
                hang_rows.append({
                    "법령명": 법령명,
                    "조문번호": display,
                    "항번호": None,
                    "항내용": None,
                    "항제개정유형": None,
                    "항제개정일자": None,
                    "호_목_목록": "\n".join(호목_lines),
                })

    # 부칙
    buches = []
    for 부칙 in root.iter("부칙단위"):
        buches.append({
            "법령명": 법령명,
            "부칙공포일자": 부칙.findtext("부칙공포일자", "").strip(),
            "부칙공포번호": 부칙.findtext("부칙공포번호", "").strip(),
            "부칙내용": 부칙.findtext("부칙내용", "").strip(),
        })

    # 개정이유
    reasons = []
    이유el = root.find(".//제개정이유")
    문내용el = root.find(".//개정문내용")
    if 이유el is not None or 문내용el is not None:
        reasons.append({
            "법령명": 법령명,
            "제개정이유": (이유el.text or "").strip() if 이유el is not None else "",
            "개정문내용": (문내용el.text or "").strip() if 문내용el is not None else "",
        })

    return basic, articles, hang_rows, buches, reasons


# ── 수집 공통 함수 ────────────────────────────────────────────────────

def collect_and_write(name: str, law_type_filter: str | None,
                      ws_basic, ws_articles, ws_hangs, ws_buches, ws_reasons):
    """법령 하나를 검색·파싱해서 각 시트에 행을 추가한다."""
    print(f"  [{name}] 검색 중...")
    results = search_law(name, law_type_filter)
    if not results:
        print(f"    => 검색 결과 없음 (건너뜀)")
        return False

    mst = results[0]["MST"]
    법종 = results[0]["법령구분"]
    print(f"    => MST: {mst} / 법종: {법종} / 수집 중...")
    xml_text = get_law_text(mst)
    b, a, h, bu, r = parse_law(xml_text)

    ws_basic.append(list(b.values()))
    for row in a:
        ws_articles.append(list(row.values()))
    for row in h:
        ws_hangs.append(list(row.values()))
    for row in bu:
        ws_buches.append(list(row.values()))
    for row in r:
        ws_reasons.append(list(row.values()))

    return True


# ── 실행부 ────────────────────────────────────────────────────────────

if __name__ == "__main__":
    wb = Workbook()
    ws_basic = wb.active
    ws_basic.title = "기본정보"
    ws_articles = wb.create_sheet("조문")
    ws_hangs = wb.create_sheet("항호목")
    ws_buches = wb.create_sheet("부칙")
    ws_reasons = wb.create_sheet("개정이유")

    ws_basic.append(["법령명_한글", "법령명_한자", "법령명_약칭", "법령ID", "법종구분",
                     "소관부처", "소관부처코드", "공포일자", "공포번호", "시행일자",
                     "제개정구분", "한글법령여부", "별표편집여부", "공포법령여부", "편장절관", "전화번호"])
    ws_articles.append(["법령명", "조문번호", "조문구분", "조문제목", "조문내용",
                        "조문시행일자", "조문변경여부", "조문참고자료"])
    ws_hangs.append(["법령명", "조문번호", "항번호", "항내용", "항제개정유형", "항제개정일자", "호_목_목록"])
    ws_buches.append(["법령명", "부칙공포일자", "부칙공포번호", "부칙내용"])
    ws_reasons.append(["법령명", "제개정이유", "개정문내용"])

    success, fail = 0, 0

    print("\n[1/3] 법률 수집")
    for name in LAW_NAMES:
        ok = collect_and_write(name, "법률", ws_basic, ws_articles, ws_hangs, ws_buches, ws_reasons)
        success += ok
        fail += not ok

    print("\n[2/3] 시행령 수집")
    for name in ENFORCEMENT_DECREE_NAMES:
        ok = collect_and_write(name, "대통령령", ws_basic, ws_articles, ws_hangs, ws_buches, ws_reasons)
        success += ok
        fail += not ok

    print("\n[3/3] 시행규칙 수집")
    for name in ENFORCEMENT_RULE_NAMES:
        ok = collect_and_write(name, "부령", ws_basic, ws_articles, ws_hangs, ws_buches, ws_reasons)
        success += ok
        fail += not ok

    # 프로젝트 루트 기준 상대경로로 저장
    out_path = Path(__file__).parent.parent / "raw" / "법령데이터.xlsx"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"\n수집 완료: 성공 {success}개 / 실패 {fail}개")
    print(f"저장 경로: {out_path}")