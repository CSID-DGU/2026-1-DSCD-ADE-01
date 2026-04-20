"""법령 원문 데이터를 parent/child 구조로 변환한다."""

import re
import csv
from pathlib import Path

import openpyxl


# ── 경로 설정 ─────────────────────────────────────────────────────────

# law_processor.py 기준: data/processors/ → data/
DATA_DIR = Path(__file__).parent.parent
INPUT_PATH = DATA_DIR / "raw" / "법령데이터.xlsx"
OUTPUT_DIR = DATA_DIR / "law_chunks"

# ── 상수 ──────────────────────────────────────────────────────────────

# 조문 삭제 패턴: "제n조 삭제", "제n조의n 삭제" 등
DELETED_RE = re.compile(r"삭제\s*[<〈]")

# 원기호(①②...) → 숫자 변환 매핑
CIRCLE_MAP = {
    "①": "1", "②": "2", "③": "3", "④": "4", "⑤": "5",
    "⑥": "6", "⑦": "7", "⑧": "8", "⑨": "9", "⑩": "10",
    "⑪": "11", "⑫": "12", "⑬": "13", "⑭": "14", "⑮": "15",
    "⑯": "16", "⑰": "17", "⑱": "18", "⑲": "19", "⑳": "20",
    "㉑": "21", "㉒": "22", "㉓": "23", "㉔": "24", "㉕": "25",
}

CIRCLE_RE = re.compile(r"[①-⑳㉑-㉟]")

PARENT_HEADER = [
    "article_key", "law_name", "law_abbr", "ministry", "enforcement_date",
    "article_no", "article_title", "article_date", "is_amended", "is_deleted",
    "parent_text", "is_article_only",
]

CHILD_HEADER = [
    "clause_key", "article_key", "law_name", "article_no",
    "paragraph_no", "child_text",
]


# ── 유틸 함수 ─────────────────────────────────────────────────────────

def circle_to_num(text: str) -> str:
    """원기호(①②...)를 숫자(1, 2...)로 변환한다."""
    if not text:
        return text
    for circle, num in CIRCLE_MAP.items():
        text = text.replace(circle, num)
    return text


def format_article_key(law_name: str, article_no: str) -> str:
    """법령명과 조문번호로 article_key를 생성한다.

    예) "주택임대차보호법", "3의2" → "주택임대차보호법_제3조의2"
    """
    if "의" in article_no:
        parts = article_no.split("의", 1)
        formatted = f"제{parts[0]}조의{parts[1]}"
    else:
        formatted = f"제{article_no}조"
    return f"{law_name}_{formatted}"


def format_clause_key(law_name: str, article_no: str, paragraph_no: str) -> str:
    """법령명, 조문번호, 항번호로 clause_key를 생성한다.

    항 있으면: "주택임대차보호법_제3조_제1항"
    항 없으면: "주택임대차보호법_제3조"
    """
    article_key = format_article_key(law_name, article_no)
    if paragraph_no:
        return f"{article_key}_제{paragraph_no}항"
    return article_key


def is_deleted_article(article_content: str) -> int:
    """조문내용에 삭제 패턴이 있으면 1, 없으면 0을 반환한다."""
    if DELETED_RE.search(article_content or ""):
        return 1
    return 0


# ── 데이터 로딩 ───────────────────────────────────────────────────────

def load_basic_info(wb: openpyxl.Workbook) -> dict:
    """기본정보 시트에서 법령별 메타데이터를 로딩한다.

    반환: {법령명: {law_abbr, ministry, enforcement_date}}
    """
    ws = wb["기본정보"]
    info = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        law_name = row[0]
        info[law_name] = {
            "law_abbr": row[2] or "",           # 법령명_약칭
            "ministry": row[5] or "",            # 소관부처
            "enforcement_date": row[9] or "",    # 시행일자
        }
    return info


def load_hang_keys(wb: openpyxl.Workbook) -> set:
    """항호목 시트에서 항이 존재하는 (법령명, 조문번호) 조합을 반환한다."""
    ws = wb["항호목"]
    keys = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        keys.add((row[0], row[1]))  # (법령명, 조문번호)
    return keys


def load_articles(wb: openpyxl.Workbook) -> list[dict]:
    """조문 시트에서 조문구분이 '조문'인 행만 로딩한다."""
    ws = wb["조문"]
    articles = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[2] != "조문":  # 조문구분 필터링 (전문 제외)
            continue
        articles.append({
            "law_name": row[0],
            "article_no": str(row[1]),
            "article_title": row[3] or "",
            "article_content": row[4] or "",
            "article_date": row[5] or "",
            "is_amended": 1 if row[6] == "Y" else 0,
        })
    return articles


def load_hang_rows(wb: openpyxl.Workbook) -> list[dict]:
    """항호목 시트에서 전체 행을 로딩한다."""
    ws = wb["항호목"]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append({
            "law_name": row[0],
            "article_no": str(row[1]),
            "paragraph_no": circle_to_num(str(row[2]).strip()) if row[2] else "",
            "hang_content": circle_to_num(row[3] or ""),
            "ho_mok": row[6] or "",
        })
    return rows


# ── 변환 함수 ─────────────────────────────────────────────────────────

def build_parent_rows(articles: list[dict], basic_info: dict, hang_keys: set) -> list[dict]:
    """조문 목록으로 parent 행을 생성한다."""
    rows = []
    for article in articles:
        law_name = article["law_name"]
        article_no = article["article_no"]
        meta = basic_info.get(law_name, {})

        article_key = format_article_key(law_name, article_no)

        # 조문제목 + 조문내용 결합
        title = article["article_title"]
        content = article["article_content"]
        parent_text = f"{title} {content}".strip() if title else content

        # 항 존재 여부
        is_article_only = 0 if (law_name, article_no) in hang_keys else 1

        rows.append({
            "article_key": article_key,
            "law_name": law_name,
            "law_abbr": meta.get("law_abbr", ""),
            "ministry": meta.get("ministry", ""),
            "enforcement_date": meta.get("enforcement_date", ""),
            "article_no": article_no,
            "article_title": title,
            "article_date": article["article_date"],
            "is_amended": article["is_amended"],
            "is_deleted": is_deleted_article(content),
            "parent_text": parent_text,
            "is_article_only": is_article_only,
        })
    return rows


def build_child_rows(hang_rows: list[dict]) -> list[dict]:
    """항호목 목록으로 child 행을 생성한다."""
    rows = []
    for hang in hang_rows:
        law_name = hang["law_name"]
        article_no = hang["article_no"]
        paragraph_no = hang["paragraph_no"]

        clause_key = format_clause_key(law_name, article_no, paragraph_no)
        article_key = format_article_key(law_name, article_no)

        # 항내용 + 호목 결합
        hang_content = hang["hang_content"]
        ho_mok = hang["ho_mok"]
        child_text = f"{hang_content}\n{ho_mok}".strip() if ho_mok else hang_content

        rows.append({
            "clause_key": clause_key,
            "article_key": article_key,
            "law_name": law_name,
            "article_no": article_no,
            "paragraph_no": int(paragraph_no) if paragraph_no else "",
            "child_text": child_text,
        })
    return rows


# ── CSV 저장 ──────────────────────────────────────────────────────────

def write_csv(path: Path, header: list[str], rows: list[dict]) -> None:
    """행 목록을 CSV 파일로 저장한다."""
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=header)
        writer.writeheader()
        writer.writerows(rows)


# ── 실행부 ────────────────────────────────────────────────────────────

def run(
    input_path: str | Path = INPUT_PATH,
    output_dir: str | Path = OUTPUT_DIR,
) -> dict:
    """법령 데이터를 parent/child CSV로 변환한다."""
    input_path = Path(input_path)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    print(f"[법령 processor] 입력: {input_path}")
    print(f"[법령 processor] 출력: {output_dir}")

    print("  엑셀 파일 로딩 중...")
    wb = openpyxl.load_workbook(input_path, read_only=True)

    basic_info = load_basic_info(wb)
    hang_keys = load_hang_keys(wb)
    articles = load_articles(wb)
    hang_rows = load_hang_rows(wb)

    print(f"  조문(조문구분=조문): {len(articles)}건")
    print(f"  항호목: {len(hang_rows)}건")

    parent_rows = build_parent_rows(articles, basic_info, hang_keys)
    child_rows = build_child_rows(hang_rows)

    # 삭제 조문 통계
    deleted_count = sum(1 for r in parent_rows if r["is_deleted"] == 1)
    amended_count = sum(1 for r in parent_rows if r["is_amended"] == 1)
    article_only_count = sum(1 for r in parent_rows if r["is_article_only"] == 1)

    parent_path = output_dir / "law_parent.csv"
    child_path = output_dir / "law_child.csv"

    write_csv(parent_path, PARENT_HEADER, parent_rows)
    write_csv(child_path, CHILD_HEADER, child_rows)

    print(f"\n=== 결과 요약 ===")
    print(f"parent: {len(parent_rows)}건 → {parent_path.name}")
    print(f"  개정 조문(is_amended=1): {amended_count}건")
    print(f"  삭제 조문(is_deleted=1): {deleted_count}건")
    print(f"  항 없는 조문(is_article_only=1): {article_only_count}건")
    print(f"child:  {len(child_rows)}건 → {child_path.name}")

    return {
        "parent_count": len(parent_rows),
        "child_count": len(child_rows),
        "deleted_count": deleted_count,
        "amended_count": amended_count,
        "article_only_count": article_only_count,
    }


if __name__ == "__main__":
    run()
